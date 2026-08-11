"""Excel-based repository implementation.

Implements repository interfaces using Excel as the persistence layer.
Excel is treated as an implementation detail, not an architectural dependency.
Future backends (SQLite, PostgreSQL) can be swapped without changing business logic.
"""

from __future__ import annotations

import json
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from domain.models import (
    Channel,
    Comment,
    RecommendationObservation,
    Video,
    VideoObservation,
)
from persistence.interfaces import (
    IChannelRepository,
    ICommentRepository,
    IObservationRepository,
    IVideoRepository,
)

logger = logging.getLogger(__name__)


class ExcelRepository(IChannelRepository, IVideoRepository, ICommentRepository, IObservationRepository):
    """Unified Excel repository implementing all persistence interfaces."""

    def __init__(self, base_path: str = "./research_data") -> None:
        self.base_path = base_path
        os.makedirs(base_path, exist_ok=True)

        # File paths for each entity type
        self.channel_file = os.path.join(base_path, "channels.xlsx")
        self.video_file = os.path.join(base_path, "videos.xlsx")
        self.comment_file = os.path.join(base_path, "comments.xlsx")
        self.video_obs_file = os.path.join(base_path, "video_observations.xlsx")
        self.rec_obs_file = os.path.join(base_path, "recommendation_observations.xlsx")

        # In-memory caches
        self._channels: Dict[str, Channel] = {}
        self._videos: Dict[str, Video] = {}
        self._comments: Dict[str, Comment] = {}
        self._video_obs: Dict[str, List[VideoObservation]] = {}
        self._rec_obs: Dict[str, List[RecommendationObservation]] = {}

        self._load_all()

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------
    def _load_all(self) -> None:
        """Load all data from Excel files into memory."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed. Excel persistence disabled.")
            return

        self._channels = self._load_sheet(self.channel_file, Channel)
        self._videos = self._load_sheet(self.video_file, Video)
        self._comments = self._load_sheet(self.comment_file, Comment)
        self._video_obs = self._load_observations(self.video_obs_file, VideoObservation)
        self._rec_obs = self._load_observations(self.rec_obs_file, RecommendationObservation)

    def _load_sheet(self, filepath: str, model_class: Any) -> Dict[str, Any]:
        """Load an Excel sheet into a dictionary keyed by canonical ID."""
        result: Dict[str, Any] = {}
        if not os.path.exists(filepath):
            return result
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            if ws is None or ws.max_row < 2:
                return result
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                # Convert JSON strings back to objects
                for key in list(row_dict.keys()):
                    if key in ("tags", "categories", "chapters", "raw_metadata", "errors",
                               "collection_run_ids", "subscriber_count_observations",
                               "video_count_observations", "view_count_observations",
                               "like_count_observations", "comment_count_observations"):
                        if isinstance(row_dict[key], str):
                            try:
                                row_dict[key] = json.loads(row_dict[key])
                            except json.JSONDecodeError:
                                row_dict[key] = []
                    # Convert datetime strings
                    if key.endswith("_at") or key in ("upload_date", "joined_date", "posted_at", "observed_at"):
                        if isinstance(row_dict[key], str):
                            try:
                                row_dict[key] = datetime.fromisoformat(row_dict[key])
                            except ValueError:
                                row_dict[key] = None
                try:
                    obj = model_class(**row_dict)
                    if hasattr(obj, "canonical_id"):
                        result[obj.canonical_id] = obj
                    elif hasattr(obj, "video_id"):
                        result.setdefault(obj.video_id, []).append(obj)
                    elif hasattr(obj, "source_video_id"):
                        result.setdefault(obj.source_video_id, []).append(obj)
                except Exception as exc:
                    logger.debug("Failed to load row into %s: %s", model_class.__name__, exc)
        except Exception as exc:
            logger.error("Failed to load %s: %s", filepath, exc)
        return result

    def _load_observations(self, filepath: str, model_class: Any) -> Dict[str, List[Any]]:
        """Load observations grouped by video_id or source_video_id."""
        result: Dict[str, List[Any]] = {}
        if not os.path.exists(filepath):
            return result
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            if ws is None or ws.max_row < 2:
                return result
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                for key in list(row_dict.keys()):
                    if key in ("raw_source_values", "raw_metadata"):
                        if isinstance(row_dict[key], str):
                            try:
                                row_dict[key] = json.loads(row_dict[key])
                            except json.JSONDecodeError:
                                row_dict[key] = {}
                    if key.endswith("_at"):
                        if isinstance(row_dict[key], str):
                            try:
                                row_dict[key] = datetime.fromisoformat(row_dict[key])
                            except ValueError:
                                row_dict[key] = None
                try:
                    obj = model_class(**row_dict)
                    vid = getattr(obj, "video_id", getattr(obj, "source_video_id", ""))
                    result.setdefault(vid, []).append(obj)
                except Exception as exc:
                    logger.debug("Failed to load observation: %s", exc)
        except Exception as exc:
            logger.error("Failed to load %s: %s", filepath, exc)
        return result

    def _save_sheet(self, filepath: str, objects: List[Any]) -> None:
        """Save a list of objects to an Excel file."""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            if not objects:
                wb.save(filepath)
                return
            # Use dict representation from first object
            first_dict = objects[0].__dict__ if hasattr(objects[0], "__dict__") else {}
            headers = list(first_dict.keys())
            ws.append(headers)
            for obj in objects:
                row_dict = obj.__dict__ if hasattr(obj, "__dict__") else {}
                row = []
                for h in headers:
                    val = row_dict.get(h)
                    if isinstance(val, (list, dict)):
                        val = json.dumps(val, default=str)
                    elif isinstance(val, datetime):
                        val = val.isoformat()
                    row.append(val)
                ws.append(row)
            wb.save(filepath)
        except Exception as exc:
            logger.error("Failed to save %s: %s", filepath, exc)

    def _persist(self) -> None:
        """Persist all in-memory data to Excel files."""
        self._save_sheet(self.channel_file, list(self._channels.values()))
        self._save_sheet(self.video_file, list(self._videos.values()))
        self._save_sheet(self.comment_file, list(self._comments.values()))
        # Flatten observations
        all_video_obs = []
        for obs_list in self._video_obs.values():
            all_video_obs.extend(obs_list)
        self._save_sheet(self.video_obs_file, all_video_obs)
        all_rec_obs = []
        for obs_list in self._rec_obs.values():
            all_rec_obs.extend(obs_list)
        self._save_sheet(self.rec_obs_file, all_rec_obs)

    # ------------------------------------------------------------------
    # IChannelRepository
    # ------------------------------------------------------------------
    def save(self, channel: Channel) -> None:
        existing = self._channels.get(channel.channel_id)
        if existing:
            # Merge: update fields, preserve first_seen_at, append collection runs
            channel.first_seen_at = existing.first_seen_at
            merged_runs = list(set(existing.collection_runs + channel.collection_runs))
            channel.collection_runs = merged_runs
        self._channels[channel.channel_id] = channel
        self._persist()

    def get_by_id(self, channel_id: str) -> Optional[Channel]:
        return self._channels.get(channel_id)

    def get_all(self) -> List[Channel]:
        return list(self._channels.values())

    def exists(self, channel_id: str) -> bool:
        return channel_id in self._channels

    # ------------------------------------------------------------------
    # IVideoRepository
    # ------------------------------------------------------------------
    def save(self, video: Video) -> None:
        existing = self._videos.get(video.video_id)
        if existing:
            video.first_seen_at = existing.first_seen_at
            merged_runs = list(set(existing.collection_runs + video.collection_runs))
            video.collection_runs = merged_runs
        self._videos[video.video_id] = video
        self._persist()

    def get_by_id(self, video_id: str) -> Optional[Video]:
        return self._videos.get(video_id)

    def get_by_channel(self, channel_id: str) -> List[Video]:
        return [v for v in self._videos.values() if v.channel_id == channel_id]

    def get_all(self) -> List[Video]:
        return list(self._videos.values())

    def exists(self, video_id: str) -> bool:
        return video_id in self._videos

    # ------------------------------------------------------------------
    # ICommentRepository
    # ------------------------------------------------------------------
    def save(self, comment: Comment) -> None:
        self._comments[comment.comment_id] = comment
        self._persist()

    def save_many(self, comments: List[Comment]) -> None:
        for c in comments:
            self._comments[c.comment_id] = c
        self._persist()

    def get_by_video(self, video_id: str) -> List[Comment]:
        return [c for c in self._comments.values() if c.video_id == video_id]

    def get_by_id(self, comment_id: str) -> Optional[Comment]:
        return self._comments.get(comment_id)

    def get_replies(self, parent_comment_id: str) -> List[Comment]:
        return [c for c in self._comments.values() if c.parent_comment_id == parent_comment_id]

    # ------------------------------------------------------------------
    # IObservationRepository
    # ------------------------------------------------------------------
    def save_video_observation(self, observation: VideoObservation) -> None:
        self._video_obs.setdefault(observation.video_id, []).append(observation)
        self._persist()

    def save_video_observations(self, observations: List[VideoObservation]) -> None:
        for obs in observations:
            self._video_obs.setdefault(obs.video_id, []).append(obs)
        self._persist()

    def get_video_observations(self, video_id: str) -> List[VideoObservation]:
        return list(self._video_obs.get(video_id, []))

    def save_recommendation_observation(self, observation: RecommendationObservation) -> None:
        self._rec_obs.setdefault(observation.source_video_id, []).append(observation)
        self._persist()

    def save_recommendation_observations(self, observations: List[RecommendationObservation]) -> None:
        for obs in observations:
            self._rec_obs.setdefault(obs.source_video_id, []).append(obs)
        self._persist()

    def get_recommendation_observations(self, source_video_id: str) -> List[RecommendationObservation]:
        return list(self._rec_obs.get(source_video_id, []))

    def get_all_recommendation_observations(self) -> List[RecommendationObservation]:
        result: List[RecommendationObservation] = []
        for obs_list in self._rec_obs.values():
            result.extend(obs_list)
        return result
