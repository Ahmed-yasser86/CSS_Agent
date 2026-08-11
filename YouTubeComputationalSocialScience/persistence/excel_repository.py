"""
Excel repository implementation for YouTube Computational Social Science research.

Provides concrete implementation of repository interfaces using Excel files
as the persistence backend, while maintaining the repository pattern for
future storage backend replacements.
"""

import os
import uuid
import pandas as pd
from typing import List, Optional, Dict, Any
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from ..domain.models import Channel, Video, Comment, Recommendation, CollectionRun, CollectionStatus, Observation
from .repository import ChannelRepository, VideoRepository, CommentRepository, RecommendationRepository, CollectionRunRepository


class ExcelRepository:
    """Base class for Excel repository implementations."""
    
    def __init__(self, base_path: str = "data/youtube"):
        """Initialize the Excel repository."""
        self.base_path = base_path
        os.makedirs(self.base_path, exist_ok=True)
    
    def _get_file_path(self, file_name: str) -> str:
        """Get the full path for an Excel file."""
        return os.path.join(self.base_path, f"{file_name}.xlsx")
    
    def _read_excel(self, file_name: str, sheet_name: str) -> Optional[pd.DataFrame]:
        """Read data from an Excel file."""
        file_path = self._get_file_path(file_name)
        if not os.path.exists(file_path):
            return None
        
        try:
            return pd.read_excel(file_path, sheet_name=sheet_name)
        except Exception:
            return None
    
    def _write_excel(self, file_name: str, sheet_name: str, df: pd.DataFrame):
        """Write data to an Excel file."""
        file_path = self._get_file_path(file_name)
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Check if file exists
        if os.path.exists(file_path):
            # Load existing workbook
            book = load_workbook(file_path)
            
            # Remove existing sheet if it exists
            if sheet_name in book.sheetnames:
                del book[sheet_name]
            
            # Create new sheet
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
        else:
            # Create new workbook
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
        
        # Write dataframe to sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.close()
    
    def _dataframe_to_model(self, df: pd.DataFrame, model_class: type, **kwargs) -> List[Any]:
        """Convert a dataframe to a list of model instances."""
        if df is None or df.empty:
            return []
        
        models = []
        for _, row in df.iterrows():
            try:
                model_data = row.to_dict()
                # Convert string representations back to original types
                for key, value in model_data.items():
                    if pd.isna(value):
                        model_data[key] = None
                    elif isinstance(value, str) and value.startswith("{") and value.endswith("}"):
                        # Try to parse JSON-like strings
                        try:
                            import json
                            model_data[key] = json.loads(value)
                        except:
                            pass
                    elif isinstance(value, str) and value.startswith("[") and value.endswith("]"):
                        # Try to parse list-like strings
                        try:
                            import ast
                            model_data[key] = ast.literal_eval(value)
                        except:
                            pass
                
                # Handle observations
                if "observations" in model_data and isinstance(model_data["observations"], str):
                    try:
                        import json
                        observations_data = json.loads(model_data["observations"])
                        observations = []
                        for obs_data in observations_data:
                            obs = Observation(**obs_data)
                            observations.append(obs)
                        model_data["observations"] = observations
                    except:
                        model_data["observations"] = []
                
                model = model_class(**model_data, **kwargs)
                models.append(model)
            except Exception as e:
                print(f"Error converting row to {model_class.__name__}: {e}")
                continue
        
        return models
    
    def _model_to_dataframe(self, models: List[Any]) -> pd.DataFrame:
        """Convert a list of model instances to a dataframe."""
        if not models:
            return pd.DataFrame()
        
        # Get all field names from the first model
        field_names = list(models[0].model_dump().keys())
        
        # Prepare data
        data = []
        for model in models:
            model_data = model.model_dump()
            
            # Convert complex objects to strings
            for key, value in model_data.items():
                if isinstance(value, list) and value and hasattr(value[0], "model_dump"):
                    # Convert list of models to list of dicts
                    model_data[key] = [item.model_dump() for item in value]
                elif hasattr(value, "model_dump"):
                    # Convert single model to dict
                    model_data[key] = value.model_dump()
                
                # Convert to string for Excel storage
                if isinstance(model_data[key], (dict, list)):
                    import json
                    model_data[key] = json.dumps(model_data[key])
            
            data.append(model_data)
        
        return pd.DataFrame(data, columns=field_names)


class ExcelChannelRepository(ExcelRepository, ChannelRepository):
    """Excel implementation of ChannelRepository."""
    
    def __init__(self, base_path: str = "data/youtube"):
        super().__init__(base_path)
    
    def save_channel(self, channel: Channel) -> bool:
        """Save a channel to the Excel repository."""
        try:
            # Convert channel to dataframe
            df = self._model_to_dataframe([channel])
            
            # Save to Excel
            self._write_excel("channels", "channels", df)
            return True
        except Exception as e:
            print(f"Error saving channel: {e}")
            return False
    
    def get_channel(self, channel_id: str) -> Optional[Channel]:
        """Retrieve a channel by its ID."""
        try:
            df = self._read_excel("channels", "channels")
            if df is None:
                return None
            
            # Filter by channel_id
            channel_df = df[df['channel_id'] == channel_id]
            if channel_df.empty:
                return None
            
            # Convert to Channel model
            channels = self._dataframe_to_model(channel_df, Channel)
            return channels[0] if channels else None
        except Exception as e:
            print(f"Error getting channel: {e}")
            return None
    
    def get_channels(self, limit: int = 100, offset: int = 0) -> List[Channel]:
        """Retrieve multiple channels with pagination."""
        try:
            df = self._read_excel("channels", "channels")
            if df is None:
                return []
            
            # Apply pagination
            paginated_df = df.iloc[offset:offset + limit]
            
            # Convert to Channel models
            return self._dataframe_to_model(paginated_df, Channel)
        except Exception as e:
            print(f"Error getting channels: {e}")
            return []
    
    def get_channel_videos(self, channel_id: str, limit: int = 100, offset: int = 0) -> List[Video]:
        """Retrieve videos for a specific channel."""
        try:
            df = self._read_excel("videos", "videos")
            if df is None:
                return []
            
            # Filter by channel_id
            videos_df = df[df['channel_id'] == channel_id]
            
            # Apply pagination
            paginated_df = videos_df.iloc[offset:offset + limit]
            
            # Convert to Video models
            return self._dataframe_to_model(paginated_df, Video)
        except Exception as e:
            print(f"Error getting channel videos: {e}")
            return []
    
    def get_channel_analytics(self, channel_id: str) -> Dict[str, Any]:
        """Retrieve analytics for a specific channel."""
        # For Excel implementation, we'll return basic analytics
        # In a real implementation, this would be stored separately
        try:
            channel = self.get_channel(channel_id)
            if not channel:
                return {}
            
            videos = self.get_channel_videos(channel_id)
            
            analytics = {
                "channel_id": channel_id,
                "video_count": len(videos),
                "subscriber_count": channel.subscriber_count,
                "total_views": sum(v.view_count for v in videos),
                "total_likes": sum(v.like_count for v in videos),
                "total_comments": sum(v.comment_count for v in videos)
            }
            
            return analytics
        except Exception as e:
            print(f"Error getting channel analytics: {e}")
            return {}
    
    def delete_channel(self, channel_id: str) -> bool:
        """Delete a channel from the repository."""
        try:
            # Read existing data
            df = self._read_excel("channels", "channels")
            if df is None:
                return False
            
            # Filter out the channel to delete
            updated_df = df[df['channel_id'] != channel_id]
            
            # Write back to Excel
            self._write_excel("channels", "channels", updated_df)
            return True
        except Exception as e:
            print(f"Error deleting channel: {e}")
            return False


class ExcelVideoRepository(ExcelRepository, VideoRepository):
    """Excel implementation of VideoRepository."""
    
    def __init__(self, base_path: str = "data/youtube"):
        super().__init__(base_path)
    
    def save_video(self, video: Video) -> bool:
        """Save a video to the Excel repository."""
        try:
            # Convert video to dataframe
            df = self._model_to_dataframe([video])
            
            # Read existing data
            existing_df = self._read_excel("videos", "videos") or pd.DataFrame()
            
            # Remove existing record if it exists
            if not existing_df.empty:
                existing_df = existing_df[existing_df['video_id'] != video.video_id]
            
            # Append new data
            updated_df = pd.concat([existing_df, df], ignore_index=True)
            
            # Save to Excel
            self._write_excel("videos", "videos", updated_df)
            return True
        except Exception as e:
            print(f"Error saving video: {e}")
            return False
    
    def get_video(self, video_id: str) -> Optional[Video]:
        """Retrieve a video by its ID."""
        try:
            df = self._read_excel("videos", "videos")
            if df is None:
                return None
            
            # Filter by video_id
            video_df = df[df['video_id'] == video_id]
            if video_df.empty:
                return None
            
            # Convert to Video model
            videos = self._dataframe_to_model(video_df, Video)
            return videos[0] if videos else None
        except Exception as e:
            print(f"Error getting video: {e}")
            return None
    
    def get_videos(self, limit: int = 100, offset: int = 0) -> List[Video]:
        """Retrieve multiple videos with pagination."""
        try:
            df = self._read_excel("videos", "videos")
            if df is None:
                return []
            
            # Apply pagination
            paginated_df = df.iloc[offset:offset + limit]
            
            # Convert to Video models
            return self._dataframe_to_model(paginated_df, Video)
        except Exception as e:
            print(f"Error getting videos: {e}")
            return []
    
    def get_video_comments(self, video_id: str, limit: int = 100, offset: int = 0) -> List[Comment]:
        """Retrieve comments for a specific video."""
        try:
            df = self._read_excel("comments", "comments")
            if df is None:
                return []
            
            # Filter by video_id
            comments_df = df[df['video_id'] == video_id]
            
            # Apply pagination
            paginated_df = comments_df.iloc[offset:offset + limit]
            
            # Convert to Comment models
            return self._dataframe_to_model(paginated_df, Comment)
        except Exception as e:
            print(f"Error getting video comments: {e}")
            return []
    
    def get_video_analytics(self, video_id: str) -> Dict[str, Any]:
        """Retrieve analytics for a specific video."""
        try:
            video = self.get_video(video_id)
            if not video:
                return {}
            
            comments = self.get_video_comments(video_id)
            
            analytics = {
                "video_id": video_id,
                "views": video.view_count,
                "likes": video.like_count,
                "comments": video.comment_count,
                "comment_count": len(comments),
                "engagement_rate": (video.like_count + video.comment_count) / video.view_count if video.view_count > 0 else 0
            }
            
            return analytics
        except Exception as e:
            print(f"Error getting video analytics: {e}")
            return {}
    
    def delete_video(self, video_id: str) -> bool:
        """Delete a video from the repository."""
        try:
            # Read existing data
            df = self._read_excel("videos", "videos")
            if df is None:
                return False
            
            # Filter out the video to delete
            updated_df = df[df['video_id'] != video_id]
            
            # Write back to Excel
            self._write_excel("videos", "videos", updated_df)
            return True
        except Exception as e:
            print(f"Error deleting video: {e}")
            return False


class ExcelCommentRepository(ExcelRepository, CommentRepository):
    """Excel implementation of CommentRepository."""
    
    def __init__(self, base_path: str = "data/youtube"):
        super().__init__(base_path)
    
    def save_comment(self, comment: Comment) -> bool:
        """Save a comment to the Excel repository."""
        try:
            # Convert comment to dataframe
            df = self._model_to_dataframe([comment])
            
            # Read existing data
            existing_df = self._read_excel("comments", "comments") or pd.DataFrame()
            
            # Remove existing record if it exists
            if not existing_df.empty:
                existing_df = existing_df[existing_df['comment_id'] != comment.comment_id]
            
            # Append new data
            updated_df = pd.concat([existing_df, df], ignore_index=True)
            
            # Save to Excel
            self._write_excel("comments", "comments", updated_df)
            return True
        except Exception as e:
            print(f"Error saving comment: {e}")
            return False
    
    def get_comment(self, comment_id: str) -> Optional[Comment]:
        """Retrieve a comment by its ID."""
        try:
            df = self._read_excel("comments", "comments")
            if df is None:
                return None
            
            # Filter by comment_id
            comment_df = df[df['comment_id'] == comment_id]
            if comment_df.empty:
                return None
            
            # Convert to Comment model
            comments = self._dataframe_to_model(comment_df, Comment)
            return comments[0] if comments else None
        except Exception as e:
            print(f"Error getting comment: {e}")
            return None
    
    def get_comments(self, limit: int = 100, offset: int = 0) -> List[Comment]:
        """Retrieve multiple comments with pagination."""
        try:
            df = self._read_excel("comments", "comments")
            if df is None:
                return []
            
            # Apply pagination
            paginated_df = df.iloc[offset:offset + limit]
            
            # Convert to Comment models
            return self._dataframe_to_model(paginated_df, Comment)
        except Exception as e:
            print(f"Error getting comments: {e}")
            return []
    
    def get_video_comments(self, video_id: str, limit: int = 100, offset: int = 0) -> List[Comment]:
        """Retrieve comments for a specific video."""
        try:
            df = self._read_excel("comments", "comments")
            if df is None:
                return []
            
            # Filter by video_id
            comments_df = df[df['video_id'] == video_id]
            
            # Apply pagination
            paginated_df = comments_df.iloc[offset:offset + limit]
            
            # Convert to Comment models
            return self._dataframe_to_model(paginated_df, Comment)
        except Exception as e:
            print(f"Error getting video comments: {e}")
            return []
    
    def get_comment_replies(self, comment_id: str, limit: int = 100, offset: int = 0) -> List[Comment]:
        """Retrieve replies for a specific comment."""
        try:
            df = self._read_excel("comments", "comments")
            if df is None:
                return []
            
            # Filter by parent_id
            replies_df = df[df['parent_id'] == comment_id]
            
            # Apply pagination
            paginated_df = replies_df.iloc[offset:offset + limit]
            
            # Convert to Comment models
            return self._dataframe_to_model(paginated_df, Comment)
        except Exception as e:
            print(f"Error getting comment replies: {e}")
            return []
    
    def get_comment_analytics(self, comment_id: str) -> Dict[str, Any]:
        """Retrieve analytics for a specific comment."""
        try:
            comment = self.get_comment(comment_id)
            if not comment:
                return {}
            
            replies = self.get_comment_replies(comment_id)
            
            analytics = {
                "comment_id": comment_id,
                "likes": comment.like_count,
                "replies": comment.reply_count,
                "reply_count": len(replies),
                "engagement_score": comment.like_count + comment.reply_count
            }
            
            return analytics
        except Exception as e:
            print(f"Error getting comment analytics: {e}")
            return {}
    
    def delete_comment(self, comment_id: str) -> bool:
        """Delete a comment from the repository."""
        try:
            # Read existing data
            df = self._read_excel("comments", "comments")
            if df is None:
                return False
            
            # Filter out the comment to delete
            updated_df = df[df['comment_id'] != comment_id]
            
            # Write back to Excel
            self._write_excel("comments", "comments", updated_df)
            return True
        except Exception as e:
            print(f"Error deleting comment: {e}")
            return False


class ExcelRecommendationRepository(ExcelRepository, RecommendationRepository):
    """Excel implementation of RecommendationRepository."""
    
    def __init__(self, base_path: str = "data/youtube"):
        super().__init__(base_path)
    
    def save_recommendation(self, recommendation: Recommendation) -> bool:
        """Save a recommendation to the Excel repository."""
        try:
            # Convert recommendation to dataframe
            df = self._model_to_dataframe([recommendation])
            
            # Read existing data
            existing_df = self._read_excel("recommendations", "recommendations") or pd.DataFrame()
            
            # Remove existing record if it exists (same source and recommended video)
            if not existing_df.empty:
                existing_df = existing_df[
                    (existing_df['source_video_id'] != recommendation.source_video_id) |
                    (existing_df['recommended_video_id'] != recommendation.recommended_video_id)
                ]
            
            # Append new data
            updated_df = pd.concat([existing_df, df], ignore_index=True)
            
            # Save to Excel
            self._write_excel("recommendations", "recommendations", updated_df)
            return True
        except Exception as e:
            print(f"Error saving recommendation: {e}")
            return False
    
    def get_recommendations_for_video(self, video_id: str, limit: int = 100, offset: int = 0) -> List[Recommendation]:
        """Retrieve recommendations for a specific video."""
        try:
            df = self._read_excel("recommendations", "recommendations")
            if df is None:
                return []
            
            # Filter by source_video_id
            recs_df = df[df['source_video_id'] == video_id]
            
            # Apply pagination
            paginated_df = recs_df.iloc[offset:offset + limit]
            
            # Convert to Recommendation models
            return self._dataframe_to_model(paginated_df, Recommendation)
        except Exception as e:
            print(f"Error getting recommendations for video: {e}")
            return []
    
    def get_recommendations_from_video(self, video_id: str, limit: int = 100, offset: int = 0) -> List[Recommendation]:
        """Retrieve recommendations from a specific video."""
        try:
            df = self._read_excel("recommendations", "recommendations")
            if df is None:
                return []
            
            # Filter by recommended_video_id
            recs_df = df[df['recommended_video_id'] == video_id]
            
            # Apply pagination
            paginated_df = recs_df.iloc[offset:offset + limit]
            
            # Convert to Recommendation models
            return self._dataframe_to_model(paginated_df, Recommendation)
        except Exception as e:
            print(f"Error getting recommendations from video: {e}")
            return []
    
    def get_video_recommendations(self, video_id: str) -> List[Recommendation]:
        """Retrieve all recommendations for a video (both for and from)."""
        try:
            df = self._read_excel("recommendations", "recommendations")
            if df is None:
                return []
            
            # Filter by both source and recommended video_id
            recs_df = df[
                (df['source_video_id'] == video_id) |
                (df['recommended_video_id'] == video_id)
            ]
            
            # Convert to Recommendation models
            return self._dataframe_to_model(recs_df, Recommendation)
        except Exception as e:
            print(f"Error getting video recommendations: {e}")
            return []
    
    def get_collection_runs_for_video(self, video_id: str) -> List[CollectionRun]:
        """Retrieve collection runs for a specific video."""
        try:
            df = self._read_excel("collection_runs", "collection_runs")
            if df is None:
                return []
            
            # Filter by source_id (video_id)
            runs_df = df[df['source_id'] == video_id]
            
            # Convert to CollectionRun models
            return self._dataframe_to_model(runs_df, CollectionRun)
        except Exception as e:
            print(f"Error getting collection runs for video: {e}")
            return []
    
    def get_recommendations_by_run(self, collection_run_id: str) -> List[Recommendation]:
        """Retrieve recommendations from a specific collection run."""
        try:
            df = self._read_excel("recommendations", "recommendations")
            if df is None:
                return []
            
            # Filter by collection_run_id
            recs_df = df[df['collection_run_id'] == collection_run_id]
            
            # Convert to Recommendation models
            return self._dataframe_to_model(recs_df, Recommendation)
        except Exception as e:
            print(f"Error getting recommendations by run: {e}")
            return []
    
    def get_recommendation_network(self, video_ids: List[str]) -> Dict[str, Any]:
        """Retrieve a recommendation network for the specified videos."""
        try:
            df = self._read_excel("recommendations", "recommendations")
            if df is None:
                return {"nodes": [], "edges": []}
            
            # Filter by video_ids
            network_df = df[
                (df['source_video_id'].isin(video_ids)) |
                (df['recommended_video_id'].isin(video_ids))
            ]
            
            # Create nodes and edges
            nodes = set()
            edges = []
            
            for _, row in network_df.iterrows():
                source = row['source_video_id']
                target = row['recommended_video_id']
                rank = row.get('recommendation_rank', 1)
                
                nodes.add(source)
                nodes.add(target)
                
                edges.append({
                    "source": source,
                    "target": target,
                    "rank": rank,
                    "observed_at": row.get('observed_at')
                })
            
            return {
                "nodes": list(nodes),
                "edges": edges
            }
        except Exception as e:
            print(f"Error getting recommendation network: {e}")
            return {"nodes": [], "edges": []}
    
    def delete_recommendation(self, source_video_id: str, recommended_video_id: str) -> bool:
        """Delete a recommendation from the repository."""
        try:
            # Read existing data
            df = self._read_excel("recommendations", "recommendations")
            if df is None:
                return False
            
            # Filter out the recommendation to delete
            updated_df = df[
                (df['source_video_id'] != source_video_id) |
                (df['recommended_video_id'] != recommended_video_id)
            ]
            
            # Write back to Excel
            self._write_excel("recommendations", "recommendations", updated_df)
            return True
        except Exception as e:
            print(f"Error deleting recommendation: {e}")
            return False